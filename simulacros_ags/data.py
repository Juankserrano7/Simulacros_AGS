import json
import os
import uuid
import unicodedata
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Tuple, Optional


import pandas as pd
import psycopg2
import streamlit as st


from .ai import generate_ai_insights
from .config import MATERIAS, MAX_UPLOAD_MB


REQUIRED_COLUMNS = ["ESTUDIANTE", "PROMEDIO PONDERADO"] + MATERIAS
OPTIONAL_NUMERIC = ["PROMEDIO SIMPLE", "DESVIACIÓN ESTÁNDAR", "PP POR MATERIA"]
COLUMN_CANONICAL_MAP = {
    "estudiante": "ESTUDIANTE",
    "grado": "GRADO",
    "lectura critica": "LECTURA CRÍTICA",
    "lectura crítica": "LECTURA CRÍTICA",
    "matematicas": "MATEMÁTICAS",
    "matemáticas": "MATEMÁTICAS",
    "sociales y ciudadanas": "SOCIALES Y CIUDADANAS",
    "ciencias naturales": "CIENCIAS NATURALES",
    "ingles": "INGLÉS",
    "inglés": "INGLÉS",
    "promedio simple": "PROMEDIO SIMPLE",
    "promedio ponderado": "PROMEDIO PONDERADO",
    "desv. estandar": "DESVIACIÓN ESTÁNDAR",
    "desviación estándar": "DESVIACIÓN ESTÁNDAR",
    "pp por materia": "PP POR MATERIA",
}


def _canonicalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Normaliza encabezados al formato esperado (acentos, mayúsculas)."""
    new_cols = []
    for col in df.columns:
        key = str(col).strip().lower()
        new_cols.append(COLUMN_CANONICAL_MAP.get(key, str(col).strip()))
    df.columns = new_cols
    return df


def _clean_student_frame(df: pd.DataFrame) -> pd.DataFrame:
    df = _canonicalize_columns(df)
    df["ESTUDIANTE"] = df["ESTUDIANTE"].astype(str).str.strip()
    df["ESTUDIANTE"] = df["ESTUDIANTE"].apply(
        lambda s: unicodedata.normalize("NFKD", s)
        .encode("ascii", "ignore")
        .decode()
        .upper()
        .replace("-", " ")
    )
    df["ESTUDIANTE"] = df["ESTUDIANTE"].str.replace(r"[^A-Z0-9 ]+", " ", regex=True)
    df["ESTUDIANTE"] = df["ESTUDIANTE"].str.replace(r"\s+", " ", regex=True).str.strip()
    # Normalización específica de D'SILVA / D SILVA a DSILVA
    df["ESTUDIANTE"] = df["ESTUDIANTE"].str.replace("D SILVA", "DSILVA", regex=False)
    df["ESTUDIANTE"] = df["ESTUDIANTE"].str.replace("D  SILVA", "DSILVA", regex=False)
    df.loc[df["ESTUDIANTE"].str.lower() == "nan", "ESTUDIANTE"] = ""

    df = df[df["ESTUDIANTE"].notna()]
    df = df[df["ESTUDIANTE"] != ""]
    df = df[~df["ESTUDIANTE"].str.upper().str.contains("PROMEDIO", na=False)]
    df = df[~df["ESTUDIANTE"].str.upper().str.contains("TOTAL", na=False)]
    df = df[~df["ESTUDIANTE"].str.upper().str.contains("MEDIA", na=False)]
    df = df.drop_duplicates(subset=["ESTUDIANTE"], keep="first")

    return df.reset_index(drop=True)


def _try_read_file(path: Path) -> pd.DataFrame:
    readers = []
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        readers = [lambda: pd.read_excel(path, skiprows=1), lambda: pd.read_excel(path)]
    else:
        readers = [lambda: pd.read_csv(path, skiprows=1), lambda: pd.read_csv(path)]

    last_exc = None
    for reader in readers:
        try:
            df = reader()
            if "ESTUDIANTE" in df.columns:
                return df
        except Exception as exc:  # noqa: BLE001
            last_exc = exc
            continue
    if last_exc:
        raise last_exc
    raise ValueError("No se pudo leer el archivo del simulacro.")


def _validate_schema(df: pd.DataFrame) -> List[str]:
    errores: List[str] = []
    missing = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        errores.append(f"Faltan las columnas requeridas: {', '.join(missing)}")

    for col in [c for c in REQUIRED_COLUMNS + OPTIONAL_NUMERIC if c not in ["ESTUDIANTE"] and c in df.columns]:
        try:
            df[col] = pd.to_numeric(df[col], errors="coerce")
            if df[col].isna().all():
                errores.append(f"La columna '{col}' no contiene valores numéricos válidos.")
        except Exception as exc:  # noqa: BLE001
            errores.append(f"No se pudo convertir la columna '{col}' a número ({exc}).")

    return errores


@st.cache_data(ttl=60)
def load_all_simulacros(promocion_id: Optional[str] = None) -> Tuple[List[Dict], Dict[str, Dict], List[str]]:
    """Carga todos los simulacros e ICFES Real pertenecientes a la promoción activa desde Supabase."""
    metadatos: List[Dict] = []
    data_map: Dict[str, Dict] = {}
    errores: List[str] = []

    if not promocion_id:
        return metadatos, data_map, ["No se especificó la promoción activa."]

    db_url = os.getenv("SUPABASE_DB_URL")
    if not db_url:
        return metadatos, data_map, ["No hay conexión configurada a Supabase (SUPABASE_DB_URL)."]

    try:
        conn = psycopg2.connect(db_url)
        try:
            with conn.cursor() as cur:
                # 1. Consultar simulacros de la promoción
                cur.execute("""
                    SELECT id, nombre, origen, estado, creado_por, creado_en, insights
                    FROM simulacros
                    WHERE promocion_id = %s
                    ORDER BY creado_en ASC;
                """, (promocion_id,))
                sim_rows = cur.fetchall()

                for s_id, s_nombre, s_origen, s_estado, s_creado_por, s_creado_en, s_insights in sim_rows:
                    meta = {
                        "id": s_id,
                        "nombre": s_nombre,
                        "origen": s_origen or "upload",
                        "estado": s_estado or "ready",
                        "creado_por": s_creado_por or "sistema",
                        "creado_en": str(s_creado_en) if s_creado_en else "",
                        "insights": s_insights if isinstance(s_insights, dict) else {},
                        "errores": [],
                        "promocion_id": promocion_id
                    }
                    metadatos.append(meta)

                    # 2. Consultar resultados de estudiantes para este simulacro
                    cur.execute("""
                        SELECT 
                            e.nombre AS "ESTUDIANTE",
                            e.grado AS "GRADO",
                            e.es_inclusion,
                            rs.lectura_critica AS "LECTURA CRÍTICA",
                            rs.matematicas AS "MATEMÁTICAS",
                            rs.sociales_ciudadanas AS "SOCIALES Y CIUDADANAS",
                            rs.ciencias_naturales AS "CIENCIAS NATURALES",
                            rs.ingles AS "INGLÉS",
                            rs.promedio_simple AS "PROMEDIO SIMPLE",
                            rs.promedio_ponderado AS "PROMEDIO PONDERADO",
                            rs.desviacion_estandar AS "DESVIACIÓN ESTÁNDAR"
                        FROM resultados_simulacro rs
                        JOIN estudiantes e ON rs.estudiante_id = e.id
                        WHERE rs.simulacro_id = %s AND e.promocion_id = %s;
                    """, (s_id, promocion_id))
                    res_rows = cur.fetchall()
                    cols = [
                        "ESTUDIANTE", "GRADO", "es_inclusion", "LECTURA CRÍTICA", "MATEMÁTICAS", 
                        "SOCIALES Y CIUDADANAS", "CIENCIAS NATURALES", "INGLÉS", 
                        "PROMEDIO SIMPLE", "PROMEDIO PONDERADO", "DESVIACIÓN ESTÁNDAR"
                    ]
                    df = pd.DataFrame(res_rows, columns=cols)
                    if not df.empty:
                        df = _clean_student_frame(df)
                    for col in cols:
                        if col not in ["ESTUDIANTE", "GRADO", "es_inclusion"] and col in df.columns:
                            df[col] = pd.to_numeric(df[col], errors="coerce")
                    data_map[s_id] = {"meta": meta, "df": df}

                # 3. Consultar si hay resultados de ICFES Real para esta promoción y agregarlo como evaluación dinamicamente
                cur.execute("""
                    SELECT 
                        e.nombre AS "ESTUDIANTE",
                        e.grado AS "GRADO",
                        e.es_inclusion,
                        rir.lectura_critica AS "LECTURA CRÍTICA",
                        rir.matematicas AS "MATEMÁTICAS",
                        rir.sociales_ciudadanas AS "SOCIALES Y CIUDADANAS",
                        rir.ciencias_naturales AS "CIENCIAS NATURALES",
                        rir.ingles AS "INGLÉS",
                        rir.puntaje_global AS "PROMEDIO PONDERADO"
                    FROM resultados_icfes_real rir
                    JOIN estudiantes e ON rir.estudiante_id = e.id
                    WHERE rir.promocion_id = %s;
                """, (promocion_id,))
                real_rows = cur.fetchall()
                if real_rows:
                    cols_real = [
                        "ESTUDIANTE", "GRADO", "es_inclusion", "LECTURA CRÍTICA", "MATEMÁTICAS",
                        "SOCIALES Y CIUDADANAS", "CIENCIAS NATURALES", "INGLÉS", "PROMEDIO PONDERADO"
                    ]
                    df_real = pd.DataFrame(real_rows, columns=cols_real)
                    df_real = _clean_student_frame(df_real)
                    for col in ["LECTURA CRÍTICA", "MATEMÁTICAS", "SOCIALES Y CIUDADANAS", "CIENCIAS NATURALES", "INGLÉS", "PROMEDIO PONDERADO"]:
                        if col in df_real.columns:
                            df_real[col] = pd.to_numeric(df_real[col], errors="coerce")

                    meta_real = {
                        "id": "icfes_real_definitivo",
                        "nombre": "🎯 ICFES Real (Definitivo)",
                        "origen": "oficial",
                        "estado": "ready",
                        "creado_por": "ICFES",
                        "creado_en": "2026-05-15 23:59:59",
                        "insights": {},
                        "errores": [],
                        "promocion_id": promocion_id
                    }
                    metadatos.append(meta_real)
                    data_map["icfes_real_definitivo"] = {"meta": meta_real, "df": df_real}

                return metadatos, data_map, errores
        finally:
            conn.close()
    except Exception as exc:  # noqa: BLE001
        return metadatos, data_map, [f"Error consultando Supabase: {exc}"]



@st.cache_data(ttl=60)
def load_icfes_real_data(promocion_id: Optional[str] = None) -> pd.DataFrame:
    """Carga los resultados oficiales del ICFES Real para la promoción dada desde Supabase."""
    if not promocion_id:
        return pd.DataFrame()
    db_url = os.getenv("SUPABASE_DB_URL")
    if not db_url:
        return pd.DataFrame()
    try:
        conn = psycopg2.connect(db_url)
        try:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT 
                        e.nombre AS "ESTUDIANTE",
                        e.es_inclusion,
                        rir.lectura_critica AS "LECTURA CRÍTICA",
                        rir.matematicas AS "MATEMÁTICAS",
                        rir.sociales_ciudadanas AS "SOCIALES Y CIUDADANAS",
                        rir.ciencias_naturales AS "CIENCIAS NATURALES",
                        rir.ingles AS "INGLÉS",
                        rir.puntaje_global AS "PROMEDIO PONDERADO"
                    FROM resultados_icfes_real rir
                    JOIN estudiantes e ON rir.estudiante_id = e.id
                    WHERE rir.promocion_id = %s;
                """, (promocion_id,))
                rows = cur.fetchall()
                cols = ["ESTUDIANTE", "es_inclusion", "LECTURA CRÍTICA", "MATEMÁTICAS", "SOCIALES Y CIUDADANAS", "CIENCIAS NATURALES", "INGLÉS", "PROMEDIO PONDERADO"]
                df = pd.DataFrame(rows, columns=cols)
                num_cols = ["LECTURA CRÍTICA", "MATEMÁTICAS", "SOCIALES Y CIUDADANAS", "CIENCIAS NATURALES", "INGLÉS", "PROMEDIO PONDERADO"]
                for col in num_cols:
                    if col in df.columns:
                        df[col] = pd.to_numeric(df[col], errors="coerce")
                return df
        finally:
            conn.close()
    except Exception:
        return pd.DataFrame()




def generate_template_bytes() -> bytes:
    """Genera la plantilla Excel en memoria y devuelve los bytes."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Simulacro"
    headers = [
        "ESTUDIANTE",
        "GRADO",
        "LECTURA CRITICA",
        "MATEMATICAS",
        "SOCIALES Y CIUDADANAS",
        "CIENCIAS NATURALES",
        "INGLES",
        "PROMEDIO SIMPLE",
        "PROMEDIO PONDERADO",
        "DESV. ESTANDAR",
        "PP POR MATERIA",
    ]
    ws.append(headers)
    estudiantes = [
        "ALVIAR MERCHAN JUAN NICOLAS",
        "BARBOSA REY MATEO",
        "CALDERON ARDILA GUIDO ARTURO",
        "CASTRO OSORIO JUAN DIEGO",
        "DELGADO ABRIL ALEJANDRO",
        "DSILVA ROSAS ALEJANDRO PABLO",
        "DURAN TORRES JUAN SEBASTIAN",
        "ESPITIA CASTRO NELSON ANDRES",
        "GARCIA ESCOBAR ESTEBAN",
        "GARCIA ZAMBRANO SAMUEL EDUARD",
        "HAZBON HERNANDEZ MANUEL FELIPE",
        "JACOME REYES ALEJANDRO",
        "MANOSALVA DURAN JUAN DIEGO",
        "MELCHIORE QUINTERO JUAN FERNANDO",
        "MUJICA ARDILA NELSON DAVID",
        "PALACIOS GOMEZ CAMILO",
        "PARRA LINARES DANIEL EDUARDO",
        "QUINTERO OROZCO SANTIAGO",
        "RANGEL CAMACHO JERONIMO",
        "REY ROJAS JUAN JOSE",
        "SALCEDO VALDIVIESO JUAN RAFAEL",
        "SANABRIA TORRES NICOLAS",
        "SERRANO MESA JERONIMO",
        "SUAREZ DURAN TOMAS",
        "TORRES ORDUZ TOMÁS",
        "TORRES RIOS EMMANUEL",
        "TORRES RODRIGUEZ ALVARO SEBASTIAN",
        "VALDERRAMA TORRES JUAN JOSE",
        "VILLAMIZAR NAVARRO JUAN JOSE",
        "VISBAL MONDRAGON NICOLAS",
    ]
    col_map = {name: get_column_letter(idx + 1) for idx, name in enumerate(headers)}
    for row_idx, est in enumerate(estudiantes, start=2):
        ws.append([est] + [""] * (len(headers) - 1))
        ws[f"{col_map['PROMEDIO SIMPLE']}{row_idx}"] = (
            f"=SUM({col_map['LECTURA CRITICA']}{row_idx}:{col_map['INGLES']}{row_idx})"
        )
        ws[f"{col_map['PP POR MATERIA']}{row_idx}"] = (
            f"=(({col_map['LECTURA CRITICA']}{row_idx}*3)"
            f"+({col_map['MATEMATICAS']}{row_idx}*3)"
            f"+({col_map['SOCIALES Y CIUDADANAS']}{row_idx}*3)"
            f"+({col_map['CIENCIAS NATURALES']}{row_idx}*3)"
            f"+({col_map['INGLES']}{row_idx}*1))/13"
        )
        ws[f"{col_map['PROMEDIO PONDERADO']}{row_idx}"] = f"={col_map['PP POR MATERIA']}{row_idx}*5"
        ws[f"{col_map['DESV. ESTANDAR']}{row_idx}"] = (
            f"=STDEV.S({col_map['LECTURA CRITICA']}{row_idx}:{col_map['INGLES']}{row_idx})"
        )

    from openpyxl.worksheet.table import Table, TableStyleInfo

    total_rows = 1 + len(estudiantes)
    table_ref = f"A1:K{total_rows}"
    table = Table(displayName="SimulacroTabla", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    ws_ins = wb.create_sheet("Instrucciones")
    ws_ins["A1"] = "Instrucciones"
    ws_ins["A2"] = "- No cambies los nombres de las columnas."
    ws_ins["A3"] = "- Ingresa un estudiante por fila."
    ws_ins["A4"] = "- Las materias y los promedios deben ser numéricos."
    ws_ins["A5"] = "- Puedes dejar PROMEDIO SIMPLE, DESV. ESTANDAR y PP POR MATERIA si no los calculas; el sistema usará PROMEDIO PONDERADO."

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()



def ingest_simulacro_excel(nombre: str, file_buffer: BytesIO, usuario: str) -> Tuple[bool, str, Dict]:
    """Valida e ingesta un nuevo simulacro cargado por Excel directamente en Supabase."""
    if not nombre or not nombre.strip():
        return False, "Debes indicar un nombre para el simulacro.", {}

    file_buffer.seek(0, 2)
    size_mb = file_buffer.tell() / (1024 * 1024)
    if size_mb > MAX_UPLOAD_MB:
        return False, f"El archivo excede el límite de {MAX_UPLOAD_MB} MB.", {}
    file_buffer.seek(0)
    header = file_buffer.read(8)
    file_buffer.seek(0)

    # Validar firma de archivo Excel (Magic Bytes)
    fname = getattr(file_buffer, "name", "").lower()
    if fname.endswith(".xlsx") and not header.startswith(b"PK\x03\x04"):
        return False, "El archivo .xlsx no posee una cabecera binaria válida de hoja de cálculo.", {}
    elif fname.endswith(".xls") and not header.startswith(b"\xd0\xcf\x11\xe0"):
        return False, "El archivo .xls no posee una cabecera binaria válida de Excel.", {}

    try:
        df_raw = pd.read_excel(file_buffer)
    except Exception as exc:  # noqa: BLE001
        return False, f"No se pudo leer el Excel: {exc}", {}

    df_raw.columns = [str(c).strip() for c in df_raw.columns]
    df = _clean_student_frame(df_raw)
    errores = _validate_schema(df)
    if errores:
        return False, "; ".join(errores), {}

    promocion_id = st.session_state.get("promocion_activa_id")
    if not promocion_id:
        return False, "No hay una promoción activa seleccionada.", {}

    db_url = os.getenv("SUPABASE_DB_URL")
    if not db_url:
        return False, "No hay conexión configurada a Supabase.", {}

    sim_id = str(uuid.uuid4())
    insights = generate_ai_insights(nombre.strip(), df, materias=MATERIAS)

    conn = psycopg2.connect(db_url)
    try:
        with conn.cursor() as cur:
            # 1. Insertar registro en tabla `simulacros`
            cur.execute("""
                INSERT INTO simulacros (id, nombre, promocion_id, origen, estado, creado_por, insights)
                VALUES (%s, %s, %s, 'upload', 'ready', %s, %s::jsonb);
            """, (sim_id, nombre.strip(), promocion_id, usuario, json.dumps(insights)))

            # 2. Insertar/obtener estudiantes e insertar sus resultados
            for _, r in df.iterrows():
                st_name = str(r["ESTUDIANTE"]).strip()
                st_grado = str(r["GRADO"]).strip() if pd.notna(r.get("GRADO")) else None

                cur.execute("""
                    INSERT INTO estudiantes (nombre, grado, promocion_id)
                    VALUES (%s, %s, %s)
                    ON CONFLICT (nombre, promocion_id) DO UPDATE SET grado = EXCLUDED.grado
                    RETURNING id;
                """, (st_name, st_grado, promocion_id))
                st_id = cur.fetchone()[0]

                def safe_num(v):
                    if pd.isna(v):
                        return None
                    try:
                        return float(v)
                    except (ValueError, TypeError):
                        return None

                cur.execute("""
                    INSERT INTO resultados_simulacro (
                        simulacro_id, estudiante_id, lectura_critica, matematicas,
                        sociales_ciudadanas, ciencias_naturales, ingles,
                        promedio_simple, promedio_ponderado, desviacion_estandar
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (simulacro_id, estudiante_id) DO UPDATE SET
                        lectura_critica = EXCLUDED.lectura_critica,
                        matematicas = EXCLUDED.matematicas,
                        sociales_ciudadanas = EXCLUDED.sociales_ciudadanas,
                        ciencias_naturales = EXCLUDED.ciencias_naturales,
                        ingles = EXCLUDED.ingles,
                        promedio_simple = EXCLUDED.promedio_simple,
                        promedio_ponderado = EXCLUDED.promedio_ponderado,
                        desviacion_estandar = EXCLUDED.desviacion_estandar;
                """, (
                    sim_id, st_id,
                    safe_num(r.get("LECTURA CRÍTICA")),
                    safe_num(r.get("MATEMÁTICAS")),
                    safe_num(r.get("SOCIALES Y CIUDADANAS")),
                    safe_num(r.get("CIENCIAS NATURALES")),
                    safe_num(r.get("INGLÉS")),
                    safe_num(r.get("PROMEDIO SIMPLE")),
                    safe_num(r.get("PROMEDIO PONDERADO")),
                    safe_num(r.get("DESVIACIÓN ESTÁNDAR"))
                ))
        conn.commit()
        st.cache_data.clear()
        return True, f"Simulacro '{nombre.strip()}' subido e ingestado correctamente en Supabase.", {"id": sim_id, "nombre": nombre.strip()}
    except Exception as exc:  # noqa: BLE001
        conn.rollback()
        return False, f"Error al procesar e insertar en Supabase: {exc}", {}
    finally:
        conn.close()


def ordenar_simulacros(data_map: Dict[str, Dict]) -> List[Dict]:
    """Convierte el mapa de datos en una lista ordenada por fecha de creación."""
    simulacros = []
    for sim_id, payload in data_map.items():
        meta = payload.get("meta", {})
        meta.setdefault("id", sim_id)
        simulacros.append(
            {
                "id": sim_id,
                "nombre": meta.get("nombre", sim_id),
                "meta": meta,
                "df": payload.get("df"),
            }
        )
    simulacros.sort(key=lambda s: s["meta"].get("creado_en", s["nombre"]))
    return simulacros


def get_or_generate_insights(sim_entry: Dict) -> Dict:
    """Devuelve los insights guardados o los genera y persiste si no existen."""
    meta = sim_entry.get("meta", {})
    insights = meta.get("insights") or {}
    if insights:
        return insights
    df = sim_entry.get("df")
    nombre = meta.get("nombre", meta.get("id", "Simulacro"))
    nuevo = generate_ai_insights(nombre, df, materias=MATERIAS)
    
    sim_id = sim_entry.get("id")
    db_url = os.getenv("SUPABASE_DB_URL")
    if db_url and sim_id:
        try:
            conn = psycopg2.connect(db_url)
            try:
                with conn.cursor() as cur:
                    cur.execute("UPDATE simulacros SET insights = %s::jsonb WHERE id = %s;", (json.dumps(nuevo), str(sim_id)))
                conn.commit()
            finally:
                conn.close()
        except Exception:
            pass

    st.cache_data.clear()
    return nuevo

