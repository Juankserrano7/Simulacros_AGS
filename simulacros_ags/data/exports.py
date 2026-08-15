"""Módulo de generación y exportación de plantillas y reportes Excel binarios."""

from io import BytesIO
from typing import Dict, List, Optional

import streamlit as st
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from ..core_utils import get_db_connection


def generate_template_bytes() -> bytes:
    """Genera la plantilla Excel en memoria y devuelve los bytes formateados."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Simulacro"
    headers = [
        "ESTUDIANTE", "GRADO", "LECTURA CRITICA", "MATEMATICAS",
        "SOCIALES Y CIUDADANAS", "CIENCIAS NATURALES", "INGLES",
        "PROMEDIO SIMPLE", "PROMEDIO PONDERADO", "DESV. ESTANDAR", "PP POR MATERIA",
    ]
    ws.append(headers)

    promocion_id = st.session_state.get("promocion_activa_id") if hasattr(st, "session_state") else None
    estudiantes_ejemplo = []
    if promocion_id:
        conn = get_db_connection()
        if conn:
            try:
                with conn.cursor() as cur:
                    cur.execute("SELECT nombre FROM estudiantes WHERE promocion_id = %s ORDER BY nombre ASC;", (promocion_id,))
                    estudiantes_ejemplo = [r[0] for r in cur.fetchall()]
            except Exception:
                pass
            finally:
                conn.close()

    if not estudiantes_ejemplo:
        estudiantes_ejemplo = [f"ESTUDIANTE {i}" for i in range(1, 11)]

    col_map = {name: get_column_letter(idx + 1) for idx, name in enumerate(headers)}
    for row_idx, est in enumerate(estudiantes_ejemplo, start=2):
        ws.append([est] + [""] * (len(headers) - 1))
        ws[f"{col_map['PROMEDIO SIMPLE']}{row_idx}"] = f"=SUM({col_map['LECTURA CRITICA']}{row_idx}:{col_map['INGLES']}{row_idx})"
        ws[f"{col_map['PP POR MATERIA']}{row_idx}"] = (
            f"=(({col_map['LECTURA CRITICA']}{row_idx}*3)"
            f"+({col_map['MATEMATICAS']}{row_idx}*3)"
            f"+({col_map['SOCIALES Y CIUDADANAS']}{row_idx}*3)"
            f"+({col_map['CIENCIAS NATURALES']}{row_idx}*3)"
            f"+({col_map['INGLES']}{row_idx}*1))/13"
        )
        ws[f"{col_map['PROMEDIO PONDERADO']}{row_idx}"] = f"={col_map['PP POR MATERIA']}{row_idx}*5"
        ws[f"{col_map['DESV. ESTANDAR']}{row_idx}"] = f"=STDEV.S({col_map['LECTURA CRITICA']}{row_idx}:{col_map['INGLES']}{row_idx})"

    table = Table(displayName="SimulacroTabla", ref=f"A1:K{1 + len(estudiantes_ejemplo)}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)

    ws_ins = wb.create_sheet("Instrucciones")
    ws_ins["A1"] = "Instrucciones"
    ws_ins["A2"] = "- No cambies los nombres de las columnas."
    ws_ins["A3"] = "- Ingresa un estudiante por fila."
    ws_ins["A4"] = "- Las materias y los promedios deben ser numéricos (0 a 100 por materia, 0 a 500 global)."
    ws_ins["A5"] = "- Puedes dejar PROMEDIO SIMPLE, DESV. ESTANDAR y PP POR MATERIA si no los calculas; el sistema usará PROMEDIO PONDERADO."

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
